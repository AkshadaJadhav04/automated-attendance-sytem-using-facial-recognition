import csv
import os
from datetime import datetime, timedelta
from io import BytesIO

def generate_csv_report(data, filename, columns):
    from flask import current_app
    reports_dir = current_app.config['REPORTS_FOLDER']
    os.makedirs(reports_dir, exist_ok=True)
    filepath = os.path.join(reports_dir, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        for row in data:
            writer.writerow(row)

    return filepath

def generate_daily_report(date=None):
    from backend.models import Attendance, get_db
    if date is None:
        date = datetime.now().date()

    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT a.date, a.time, s.roll_number, s.name, s.class_division,
               sub.subject_name, a.lecture_number, a.status, a.confidence_score
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        JOIN subjects sub ON a.subject_id = sub.id
        WHERE a.date = %s
        ORDER BY sub.subject_name, s.roll_number
    """, (date,))
    records = cursor.fetchall()
    cursor.close()

    columns = ['Date', 'Time', 'Roll Number', 'Name', 'Class', 'Subject', 'Lecture', 'Status', 'Confidence']
    data = [[
        r['date'].strftime('%Y-%m-%d') if hasattr(r['date'], 'strftime') else r['date'],
        r['time'].strftime('%H:%M:%S') if hasattr(r['time'], 'strftime') else r['time'],
        r['roll_number'], r['name'], r['class_division'],
        r['subject_name'], r['lecture_number'], r['status'],
        f"{r['confidence_score']:.2f}%" if r['confidence_score'] else 'N/A'
    ] for r in records]

    filename = f"daily_report_{date}.csv"
    return generate_csv_report(data, filename, columns)

def generate_monthly_report(year, month):
    from backend.models import Attendance, get_db
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date()
    else:
        end_date = datetime(year, month + 1, 1).date()

    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT s.roll_number, s.name, s.class_division,
               COUNT(DISTINCT CONCAT(a.subject_id, '_', a.date, '_', a.lecture_number)) as total_classes,
               SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present_count,
               ROUND(
                   SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) /
                   NULLIF(COUNT(DISTINCT CONCAT(a.subject_id, '_', a.date, '_', a.lecture_number)), 0) * 100, 2
               ) as percentage
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id
            AND a.date >= %s AND a.date < %s
        GROUP BY s.id, s.roll_number, s.name, s.class_division
        ORDER BY s.roll_number
    """, (start_date, end_date))
    records = cursor.fetchall()
    cursor.close()

    columns = ['Roll Number', 'Name', 'Class', 'Total Classes', 'Present', 'Percentage']
    data = [[r['roll_number'], r['name'], r['class_division'],
             r['total_classes'], r['present_count'], f"{r['percentage']}%"] for r in records]

    filename = f"monthly_report_{year}_{month:02d}.csv"
    return generate_csv_report(data, filename, columns)

def generate_student_report(student_id):
    from backend.models import Attendance
    records = Attendance.get_by_student(student_id)
    columns = ['Date', 'Time', 'Subject', 'Lecture', 'Status', 'Confidence']
    data = [[
        r['date'].strftime('%Y-%m-%d') if hasattr(r['date'], 'strftime') else r['date'],
        r['time'].strftime('%H:%M:%S') if hasattr(r['time'], 'strftime') else r['time'],
        r['subject_name'], r['lecture_number'], r['status'],
        f"{r['confidence_score']:.2f}%" if r['confidence_score'] else 'N/A'
    ] for r in records]

    filename = f"student_report_{student_id}.csv"
    return generate_csv_report(data, filename, columns)

def generate_subject_report(subject_id, start_date=None, end_date=None):
    from backend.models import Attendance
    records = Attendance.get_by_subject(subject_id, start_date, end_date)
    columns = ['Date', 'Time', 'Roll Number', 'Name', 'Class', 'Status', 'Confidence']
    data = [[
        r['date'].strftime('%Y-%m-%d') if hasattr(r['date'], 'strftime') else r['date'],
        r['time'].strftime('%H:%M:%S') if hasattr(r['time'], 'strftime') else r['time'],
        r['roll_number'], r['name'], r['class_division'],
        r['status'],
        f"{r['confidence_score']:.2f}%" if r['confidence_score'] else 'N/A'
    ] for r in records]

    filename = f"subject_report_{subject_id}.csv"
    return generate_csv_report(data, filename, columns)

def generate_excel_report(data, columns, filename):
    try:
        import openpyxl
        from flask import current_app

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        reports_dir = current_app.config['REPORTS_FOLDER']
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        wb.save(filepath)
        return filepath
    except ImportError:
        return None

def generate_pdf_report(data, columns, filename):
    try:
        from fpdf import FPDF
        from flask import current_app

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(200, 10, "Attendance Report", ln=True, align="C")
        pdf.ln(10)

        pdf.set_font("Arial", "B", 10)
        for col in columns:
            pdf.cell(30, 10, col, 1)
        pdf.ln()

        pdf.set_font("Arial", "", 8)
        for row in data:
            for item in row:
                pdf.cell(30, 8, str(item)[:20], 1)
            pdf.ln()

        reports_dir = current_app.config['REPORTS_FOLDER']
        os.makedirs(reports_dir, exist_ok=True)
        filepath = os.path.join(reports_dir, filename)
        pdf.output(filepath)
        return filepath
    except ImportError:
        return None
